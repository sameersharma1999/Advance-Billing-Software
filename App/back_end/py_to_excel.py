from openpyxl import load_workbook
import random


class PyToPdf:

    @staticmethod
    def upload_to_excel(mainlist, transportlist, customerlist, pricelist):
        wb = load_workbook(filename="..\\back_end\\bs.xlsx")

        ws = wb['Sheet1']

        customer_list = customerlist
        # customer_list = ['Sameer', '16/250 sarabha nagar Extn.', '123456789 RT 0001', '1234 5678 9123',
        #                  'JMXPS4639K', '9779478123']
        ws.cell(row=4, column=10).value = pricelist[5]
        ws.cell(row=12, column=3).value = str(customer_list[0][0] + customer_list[0][1])
        ws.cell(row=13, column=3).value = customer_list[0][2]
        ws.cell(row=14, column=3).value = customer_list[0][3]
        ws.cell(row=15, column=3).value = customer_list[0][4]
        ws.cell(row=16, column=3).value = customer_list[0][5]
        ws.cell(row=17, column=3).value = customer_list[0][6]
        if customer_list[1] == "" or customer_list[1] is None:
            temp = 0
        else:
            temp = 1
        ws.cell(row=12, column=7).value = str(customer_list[temp][0] + customer_list[temp][1])
        ws.cell(row=13, column=7).value = customer_list[temp][1]
        ws.cell(row=14, column=7).value = customer_list[temp][2]
        ws.cell(row=15, column=7).value = customer_list[temp][3]
        ws.cell(row=16, column=7).value = customer_list[temp][4]
        ws.cell(row=17, column=7).value = customer_list[temp][5]

        item_list = mainlist

        # item_list = [['1', '12345', 'mobile phone', '#1001', '2', '10000', '2.5', '2.5', '12'],
        #              ['2', '12312', 'something1', '#1002', '4', '100', '2.5', '2.5', '2'],
        #              ['3', '12333', 'something2', '#1003', '8', '2000', '2.5', '2.5', '4'],
        #              ['4', '12399', 'something3', '#1004', '1', '3000', '2.5', '2.5', '13'],
        #              ['5', '12323', 'something4', '#1005', '0', '9000', '2.5', '2.5', '17']]

        ws.cell(row=len(item_list) + 24, column=9).value = 'Total  ' + str(pricelist[0])
        ws.cell(row=len(item_list) + 25, column=9).value = 'CGST  ' + str(pricelist[1])
        ws.cell(row=len(item_list) + 26, column=9).value = 'IGST  ' + str(pricelist[3])
        ws.cell(row=len(item_list) + 27, column=9).value = 'Grand Total  ' + str(pricelist[4])

        ws.cell(row=len(item_list) + 30, column=1).value = 'T/C'
        ws.cell(row=len(item_list) + 31, column=1).value = '1. The shipping cost needs to be beared by the seller'
        ws.cell(row=len(item_list) + 32,
                column=1).value = '2. The seller is not responsible for any damage that happens during the transit'
        ws.cell(row=len(item_list) + 33, column=1).value = '3. If invoice has not been paid in 5 days after due date,'
        ws.cell(row=len(item_list) + 34,
                column=1).value = '  a tax of 10% of total value is applied to each day of delay'

        ws.cell(row=len(item_list) + 36, column=1).value = 'TRANSPORTATION'
        ws.cell(row=len(item_list) + 37, column=1).value = 'Name: ' + str(transportlist[0])
        ws.cell(row=len(item_list) + 38, column=1).value = 'Address: ' + str(transportlist[1])
        ws.cell(row=len(item_list) + 39, column=1).value = 'Vehicle: ' + str(transportlist[2])
        # ws.cell(row=len(item_list) + 40, column=1).value = 'GST No: ' + str(transportlist[3])

        ws.cell(row=len(item_list) + 44, column=7).value = 'Proprietor Signature'

        row = 0
        t = 1
        for i in item_list:
            ws.cell(row=22 + row, column=1).value = t
            ws.cell(row=22 + row, column=2).value = i[0]
            ws.cell(row=22 + row, column=3).value = i[1]
            ws.cell(row=22 + row, column=4).value = i[2]
            ws.cell(row=22 + row, column=5).value = i[3]
            ws.cell(row=22 + row, column=6).value = i[4]
            ws.cell(row=22 + row, column=7).value = i[6]
            ws.cell(row=22 + row, column=8).value = i[7]
            ws.cell(row=22 + row, column=9).value = i[8]
            ws.cell(row=22 + row, column=10).value = i[5]
            ws.cell(row=22 + row, column=11).value = i[9]

            t += 1

            row = row + 1

        name = str(random.randint(1, 100000000))
        wb.save("..\\back_end\\" + name + '.xlsx')
        print("Done")
        return name + '.xlsx'

